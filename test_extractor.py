from openpyxl import load_workbook

from extractor import read_agenda


# setting up testing files
wb1 = load_workbook('test_example_1.xlsx', data_only=True)
worksheet_1 = wb1.worksheets[0]

wb2 = load_workbook('test_example_2.xlsx', data_only=True)
worksheet_2 = wb2.worksheets[0]

wb3 = load_workbook('test_example_3.xlsx', data_only=True)
worksheet_3 = wb3.worksheets[0]


class TestReadAgendaEventsCount:
    """
    checking if reading algorithm returns all the events
    """
    def test_worksheet_1(self):
        assert len(read_agenda(worksheet_1)) == 16

    def test_worksheet_2(self):
        assert len(read_agenda(worksheet_2)) == 15

    def test_worksheet_3(self):
        assert len(read_agenda(worksheet_3)) == 18
